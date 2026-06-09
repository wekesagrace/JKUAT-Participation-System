from flask_mail import Mail, Message
from flask import (Flask, render_template, request, redirect,
                   url_for, flash, session, send_file, jsonify)
from models import db, Feedback, User, Event
from nlp import analyze_sentiment, detect_topic, analyze_topics_lda
from sms_handler import initialize_sms, send_acknowledgement
from functools import wraps
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table,
                                 TableStyle, Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import qrcode
import os
import base64

app = Flask(__name__)

# ── Configuration ──────────────────────────────────────────
app.config['SECRET_KEY']                  = 'sautiyetu2026secretkey'
app.config['SQLALCHEMY_DATABASE_URI']     = 'sqlite:///sautiyetu.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Email Configuration ────────────────────────────────────
app.config['MAIL_SERVER']   = 'smtp.gmail.com'
app.config['MAIL_PORT']     = 587
app.config['MAIL_USE_TLS']  = True
app.config['MAIL_USERNAME'] = 'sautiyetu123@gmail.com'  # your Gmail
app.config['MAIL_PASSWORD'] = 'xqzsfodnfebvycua'      # app password
app.config['MAIL_DEFAULT_SENDER'] = 'your.email@gmail.com'


mail = Mail(app)

# Africa's Talking SMS
AT_USERNAME = 'sandbox'
AT_API_KEY  = '1xhyfbXd0'  

# Alert threshold — alert when negative % exceeds this
ALERT_THRESHOLD = 40

db.init_app(app)

with app.app_context():
    db.create_all()
    # Create default admin if none exists
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin',
                     password='admin123', role='admin')
        db.session.add(admin)
        db.session.commit()


# ── Login decorator ────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'admin' not in session:
            flash('Please login to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def send_alert_email(negative_pct, total, negative):
    """Send email alert when negative feedback is high"""
    try:
        msg = Message(
            subject='SautiYetu Alert — High Negative Feedback Detected',
            recipients=['your.email@gmail.com']  # where to send alert
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif; max-width:600px;
                    margin:0 auto; padding:20px;">

            <div style="background:#0B3D2E; padding:20px;
                        border-radius:12px 12px 0 0; text-align:center;">
                <h2 style="color:#1ABC9C; margin:0;">
                    SautiYetu Alert
                </h2>
                <p style="color:rgba(255,255,255,0.7); margin:8px 0 0;">
                    Automated Governance Notification
                </p>
            </div>

            <div style="background:#fff; padding:28px;
                        border:1px solid #e5e7eb;
                        border-radius:0 0 12px 12px;">

                <div style="background:#fef2f2; border:1px solid #fca5a5;
                            border-radius:10px; padding:16px;
                            margin-bottom:20px; text-align:center;">
                    <div style="font-size:36px; font-weight:800;
                                color:#ef4444;">
                        {negative_pct}%
                    </div>
                    <div style="font-size:14px; color:#991b1b;
                                font-weight:600;">
                        Negative Feedback Rate
                    </div>
                    <div style="font-size:12px; color:#888;
                                margin-top:4px;">
                        Above the 40% alert threshold
                    </div>
                </div>

                <table style="width:100%; border-collapse:collapse;
                              margin-bottom:20px;">
                    <tr style="background:#f9fafb;">
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#555; font-weight:600;">
                            Total Feedback
                        </td>
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#0B3D2E; font-weight:700;">
                            {total} submissions
                        </td>
                    </tr>
                    <tr>
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#555; font-weight:600;">
                            Negative Count
                        </td>
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#ef4444; font-weight:700;">
                            {negative} submissions
                        </td>
                    </tr>
                    <tr style="background:#f9fafb;">
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#555; font-weight:600;">
                            Alert Time
                        </td>
                        <td style="padding:10px 14px; font-size:13px;
                                   color:#0B3D2E; font-weight:700;">
                            {datetime.now().strftime('%d %B %Y at %I:%M %p')}
                        </td>
                    </tr>
                </table>

                <p style="font-size:13px; color:#555;
                          line-height:1.7; margin-bottom:20px;">
                    The SautiYetu platform has detected that negative
                    feedback has exceeded the 40% threshold.
                    Please review recent submissions and take
                    appropriate action.
                </p>

                <div style="text-align:center;">
                    <a href="http://127.0.0.1:5000/dashboard"
                       style="background:#0B3D2E; color:#fff;
                              border-radius:10px; padding:12px 28px;
                              text-decoration:none; font-size:14px;
                              font-weight:700; display:inline-block;">
                        View Dashboard
                    </a>
                </div>

            </div>

            <p style="text-align:center; font-size:11px;
                      color:#aaa; margin-top:16px;">
                This is an automated alert from SautiYetu
                Civic Platform · JKUAT · 2026
            </p>
        </div>
        """
        mail.send(msg)
        print("Alert email sent successfully!")
    except Exception as e:
        print(f"Email alert failed: {e}")


# ══════════════════════════════════════════════════════════
# PUBLIC ROUTES
# ══════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/feedback')
def feedback():
    return render_template('feedback.html')


@app.route('/submit', methods=['POST'])
def submit():
    name    = request.form.get('name', '').strip()
    message = request.form.get('message', '').strip()

    if not message:
        flash('Please enter your feedback before submitting.', 'warning')
        return redirect(url_for('feedback'))

    sentiment, keywords = analyze_sentiment(message)
    topic               = detect_topic(message)

    new_feedback = Feedback(
        name      = name if name else 'Anonymous',
        message   = message,
        channel   = 'web',
        sentiment = sentiment,
        keywords  = keywords,
        topic     = topic
    )
    db.session.add(new_feedback)
    db.session.commit()

    # Alert check — if negative feedback is high notify in session
    all_fb       = Feedback.query.all()
    total        = len(all_fb)
    negative     = sum(1 for f in all_fb if f.sentiment == 'Negative')
    negative_pct = round(negative / total * 100) if total > 0 else 0

    if total > 0 and negative_pct >= ALERT_THRESHOLD:
        session['alert'] = (
           f'Alert: Negative feedback has reached '
           f'{negative_pct}% — above the {ALERT_THRESHOLD}% threshold.'
    )
    # Send email alert
    send_alert_email(negative_pct, total, negative)

    flash('Thank you! Your feedback has been received and analysed.', 'success')
    return redirect(url_for('feedback'))


@app.route('/events')
def events():
    category   = request.args.get('category', None)
    if category:
        all_events = Event.query.filter_by(
            category=category, is_active=True
        ).order_by(Event.event_date).all()
    else:
        all_events = Event.query.filter_by(
            is_active=True
        ).order_by(Event.event_date).all()
    return render_template('events.html', events=all_events)


@app.route('/transcribe', methods=['GET', 'POST'])
def transcribe():
    transcription = None
    if request.method == 'POST':
        audio_file = request.files.get('audio')
        language   = request.form.get('language', 'auto')

        if not audio_file:
            flash('Please upload an audio file.', 'warning')
            return redirect(url_for('transcribe'))

        # Create uploads folder if it does not exist
        upload_folder = os.path.join(os.path.dirname(__file__), 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = secure_filename(audio_file.filename)
        filepath = os.path.join(upload_folder, filename)
        audio_file.save(filepath)

        try:
            import whisper
            model  = whisper.load_model('base')
            if language == 'auto':
                result = model.transcribe(filepath)
            else:
                result = model.transcribe(filepath, language=language)
            transcription = result['text'].strip()
            flash('Audio transcribed successfully!', 'success')
        except Exception as e:
            flash(f'Transcription error: {str(e)}', 'danger')
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

    return render_template('transcribe.html',
                           transcription=transcription)

# ══════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'admin' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user     = User.query.filter_by(
                       username=username, password=password).first()
        if user:
            session['admin'] = username
            flash(f'Welcome back, {username}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Incorrect username or password.', 'danger')
    return render_template('login.html')


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        secret   = request.form.get('secret_answer', '').strip().lower()
        new_pass = request.form.get('new_password', '').strip()
        confirm  = request.form.get('confirm_password', '').strip()

        if secret not in ['jkuat', 'jomo kenyatta']:
            flash('Security answer incorrect.', 'danger')
            return redirect(url_for('forgot_password'))
        if new_pass != confirm:
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('forgot_password'))
        if len(new_pass) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return redirect(url_for('forgot_password'))

        user = User.query.filter_by(username=username).first()
        if not user:
            flash('Username not found.', 'danger')
            return redirect(url_for('forgot_password'))

        user.password = new_pass
        db.session.commit()
        flash('Password reset successfully! You can now login.', 'success')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')


@app.route('/logout')
def logout():
    session.pop('admin', None)
    session.pop('alert', None)
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════
# ADMIN ROUTES
# ══════════════════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    feedbacks    = Feedback.query.order_by(
                       Feedback.submitted_at.desc()).all()
    total        = len(feedbacks)
    positive     = sum(1 for f in feedbacks if f.sentiment == 'Positive')
    negative     = sum(1 for f in feedbacks if f.sentiment == 'Negative')
    neutral      = sum(1 for f in feedbacks if f.sentiment == 'Neutral')
    web          = sum(1 for f in feedbacks if f.channel == 'web')
    sms          = sum(1 for f in feedbacks if f.channel == 'sms')
    assisted     = sum(1 for f in feedbacks if f.channel == 'assisted')
    all_messages = [f.message for f in feedbacks]
    topic_counts = analyze_topics_lda(all_messages)

    # Check alert
    alert_msg = None
    if total > 0 and (negative / total * 100) >= ALERT_THRESHOLD:
        alert_msg = (
            f'Negative sentiment has reached '
            f'{round(negative/total*100)}% '
            f'— above the {ALERT_THRESHOLD}% threshold. '
            f'Review recent submissions immediately.'
        )

    return render_template('dashboard.html',
        feedbacks    = feedbacks,
        total        = total,
        positive     = positive,
        negative     = negative,
        neutral      = neutral,
        web          = web,
        sms          = sms,
        assisted     = assisted,
        topic_counts = topic_counts,
        alert_msg    = alert_msg
    )


@app.route('/insights')
@login_required
def insights():
    feedbacks    = Feedback.query.order_by(
                       Feedback.submitted_at.desc()).all()
    total        = len(feedbacks)
    positive     = sum(1 for f in feedbacks if f.sentiment == 'Positive')
    negative     = sum(1 for f in feedbacks if f.sentiment == 'Negative')
    neutral      = sum(1 for f in feedbacks if f.sentiment == 'Neutral')
    all_messages = [f.message for f in feedbacks]
    topic_counts = analyze_topics_lda(all_messages)

    # Daily trend — last 7 days
    from collections import defaultdict
    daily = defaultdict(lambda: {'Positive': 0,
                                  'Negative': 0, 'Neutral': 0})
    for f in feedbacks:
        if f.submitted_at:
            day = f.submitted_at.strftime('%d %b')
            daily[day][f.sentiment] += 1

    return render_template('insights.html',
        feedbacks    = feedbacks,
        total        = total,
        positive     = positive,
        negative     = negative,
        neutral      = neutral,
        topic_counts = topic_counts,
        daily        = dict(daily)
    )


@app.route('/events/add', methods=['GET', 'POST'])
@login_required
def add_event():
    if request.method == 'POST':
        title       = request.form.get('title')
        description = request.form.get('description')
        location    = request.form.get('location')
        event_date  = request.form.get('event_date')
        event_time  = request.form.get('event_time')
        category    = request.form.get('category', 'General')

        new_event = Event(
            title       = title,
            description = description,
            location    = location,
            event_date  = event_date,
            event_time  = event_time,
            category    = category
        )
        db.session.add(new_event)
        db.session.commit()
        flash('Event posted successfully!', 'success')
        return redirect(url_for('events'))
    return render_template('add_event.html')


@app.route('/events/delete/<int:id>')
@login_required
def delete_event(id):
    event = Event.query.get_or_404(id)
    event.is_active = False
    db.session.commit()
    flash('Event removed.', 'info')
    return redirect(url_for('events'))


@app.route('/events/qr/<int:id>')
def event_qr(id):
    """Generate QR code for an event"""
    event     = Event.query.get_or_404(id)
    event_url = (
        f"Event: {event.title}\n"
        f"Date: {event.event_date}\n"
        f"Time: {event.event_time}\n"
        f"Venue: {event.location}\n"
        f"Submit feedback: http://127.0.0.1:5000/feedback"
    )
    qr     = qrcode.QRCode(box_size=8, border=4)
    qr.add_data(event_url)
    qr.make(fit=True)
    img    = qr.make_image(fill_color='#0B3D2E',
                            back_color='white')
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_b64 = base64.b64encode(buffer.read()).decode('utf-8')
    return render_template('event_qr.html',
                           event=event, qr_code=img_b64)


# ── Live dashboard data API ────────────────────────────────
@app.route('/api/live-data')
@login_required
def live_data():
    feedbacks = Feedback.query.order_by(
                    Feedback.submitted_at.desc()).all()
    total     = len(feedbacks)
    positive  = sum(1 for f in feedbacks if f.sentiment == 'Positive')
    negative  = sum(1 for f in feedbacks if f.sentiment == 'Negative')
    neutral   = sum(1 for f in feedbacks if f.sentiment == 'Neutral')
    web       = sum(1 for f in feedbacks if f.channel == 'web')
    sms       = sum(1 for f in feedbacks if f.channel == 'sms')
    assisted  = sum(1 for f in feedbacks if f.channel == 'assisted')
    alert     = (total > 0 and
                 (negative / total * 100) >= ALERT_THRESHOLD)
    return jsonify({
        'total':    total,
        'positive': positive,
        'negative': negative,
        'neutral':  neutral,
        'web':      web,
        'sms':      sms,
        'assisted': assisted,
        'alert':    alert
    })


# ── Export routes ──────────────────────────────────────────
@app.route('/export/pdf')
@login_required
def export_pdf():
    feedbacks = Feedback.query.order_by(
                    Feedback.submitted_at.desc()).all()
    buffer    = BytesIO()
    doc       = SimpleDocTemplate(buffer, pagesize=A4,
                    rightMargin=2*cm, leftMargin=2*cm,
                    topMargin=2*cm, bottomMargin=2*cm)
    styles   = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        'SautiYetu — Feedback Report', styles['Title']))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %B %Y %I:%M %p')}",
        styles['Normal']))
    elements.append(Spacer(1, 0.4*cm))

    total    = len(feedbacks)
    positive = sum(1 for f in feedbacks if f.sentiment == 'Positive')
    negative = sum(1 for f in feedbacks if f.sentiment == 'Negative')
    neutral  = sum(1 for f in feedbacks if f.sentiment == 'Neutral')

    summary = Table(
        [['Total', 'Positive', 'Neutral', 'Negative'],
         [str(total), str(positive), str(neutral), str(negative)]],
        colWidths=[4*cm, 4*cm, 4*cm, 4*cm]
    )
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B3D2E')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#E0F7F4')),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(summary)
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph('All Feedback', styles['Heading2']))

    rows = [['#', 'Name', 'Message', 'Sentiment', 'Topic',
             'Channel', 'Date']]
    for i, f in enumerate(feedbacks, 1):
        rows.append([
            str(i),
            (f.name or 'Anonymous')[:15],
            (f.message[:55] + '...'
             if len(f.message) > 55 else f.message),
            f.sentiment,
            (f.topic or 'General'),
            f.channel,
            f.submitted_at.strftime('%d %b %Y')
            if f.submitted_at else 'N/A'
        ])

    tbl = Table(rows,
                colWidths=[0.8*cm, 2.5*cm, 6*cm,
                            2*cm, 2.5*cm, 1.8*cm, 2.4*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#0B3D2E')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('GRID',          (0,0), (-1,-1), 0.3, colors.grey),
        ('ROWBACKGROUNDS',(0,1), (-1,-1),
         [colors.white, colors.HexColor('#F0F7F5')]),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'SautiYetu_Report_'
                     f'{datetime.now().strftime("%Y%m%d")}.pdf')


@app.route('/export/excel')
@login_required
def export_excel():
    feedbacks = Feedback.query.order_by(
                    Feedback.submitted_at.desc()).all()
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Feedback Data'

    hdr_fill  = PatternFill('solid', fgColor='0B3D2E')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_align = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)

    headers = ['#', 'Name', 'Message', 'Sentiment',
               'Topic', 'Keywords', 'Channel', 'Date']
    for col, h in enumerate(headers, 1):
        c           = ws1.cell(row=1, column=col, value=h)
        c.fill      = hdr_fill
        c.font      = hdr_font
        c.alignment = hdr_align

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 30
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 18

    pos_fill = PatternFill('solid', fgColor='C8E6C9')
    neg_fill = PatternFill('solid', fgColor='FFCDD2')
    neu_fill = PatternFill('solid', fgColor='FFF9C4')

    for i, f in enumerate(feedbacks, 2):
        ws1.cell(row=i, column=1, value=i-1)
        ws1.cell(row=i, column=2, value=f.name or 'Anonymous')
        c3 = ws1.cell(row=i, column=3, value=f.message)
        c3.alignment = Alignment(wrap_text=True)
        c4 = ws1.cell(row=i, column=4, value=f.sentiment)
        if f.sentiment == 'Positive':
            c4.fill = pos_fill
        elif f.sentiment == 'Negative':
            c4.fill = neg_fill
        else:
            c4.fill = neu_fill
        ws1.cell(row=i, column=5, value=f.topic or 'General')
        ws1.cell(row=i, column=6, value=f.keywords or '')
        ws1.cell(row=i, column=7, value=f.channel)
        ws1.cell(row=i, column=8,
                 value=f.submitted_at.strftime('%d %b %Y %H:%M')
                 if f.submitted_at else 'N/A')

    # Summary sheet
    ws2       = wb.create_sheet('Summary')
    total     = len(feedbacks)
    positive  = sum(1 for f in feedbacks if f.sentiment == 'Positive')
    negative  = sum(1 for f in feedbacks if f.sentiment == 'Negative')
    neutral   = sum(1 for f in feedbacks if f.sentiment == 'Neutral')
    ws2['A1'] = 'SautiYetu Feedback Summary'
    ws2['A1'].font = Font(bold=True, size=14, color='0B3D2E')
    ws2['A2'] = f'Generated: {datetime.now().strftime("%d %B %Y")}'
    for row in [
        ['', ''], ['SENTIMENT', ''],
        ['Total', total], ['Positive', positive],
        ['Neutral', neutral], ['Negative', negative],
        ['', ''], ['CHANNEL', ''],
        ['Web', sum(1 for f in feedbacks if f.channel == 'web')],
        ['SMS', sum(1 for f in feedbacks if f.channel == 'sms')],
        ['Assisted',
         sum(1 for f in feedbacks if f.channel == 'assisted')],
    ]:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer,
        mimetype='application/vnd.openxmlformats-officedocument'
                 '.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'SautiYetu_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ── SMS webhook ────────────────────────────────────────────
@app.route('/sms/incoming', methods=['POST'])
def sms_incoming():
    phone   = request.form.get('from', 'Unknown')
    message = request.form.get('text', '').strip()
    if not message:
        return 'OK', 200

    sentiment, keywords = analyze_sentiment(message)
    topic               = detect_topic(message)

    db.session.add(Feedback(
        name      = f'SMS: {phone}',
        message   = message,
        channel   = 'sms',
        sentiment = sentiment,
        keywords  = keywords,
        topic     = topic
    ))
    db.session.commit()

    try:
        svc = initialize_sms(AT_USERNAME, AT_API_KEY)
        send_acknowledgement(svc, phone)
    except Exception as e:
        print(f'SMS ACK failed: {e}')
    return 'OK', 200


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')